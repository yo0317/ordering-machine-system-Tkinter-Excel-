# ＿＿author＿＿：CHANG,YOU-HSUAN/yo
from tkinter import ttk
from tkinter import messagebox
import tkinter as tk
from tkinter import *
from turtle import heading
from PIL import ImageTk, Image
from tkcalendar import Calendar
from tkinter.ttk import Notebook, Style
import matplotlib.pyplot as plt         # 匯入 matplotlib 的 pyplot 類別,並設定為 plt
import sys
import datetime as dt
from openpyxl import Workbook
import pandas as pd
import os
from openpyxl import load_workbook


if sys.platform.startswith("linux"):  # could be "linux", "linux2", "linux3", ...
    print("linux")  # linux
elif sys.platform == "darwin":  # MAC OS X
    from matplotlib.font_manager import FontProperties      # 中文字體
    plt.rcParams['font.sans-serif'] = 'Arial Unicode MS'
    plt.rcParams['axes.unicode_minus'] = False
    wordColor = "black"
    font0 = ("Helvetica", 8)
    font1 = ("Helvetica", 16)
    font2 = ("Helvetica", 32)
    font3 = ("Helvetica", 64)
elif sys.platform == "win32":
    # Windows (either 32-bit or 64-bit)
    plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei']  # 換成中文的字體
    plt.rcParams['axes.unicode_minus'] = False                # 步驟二（解決座標軸負數的負號顯示問題）
    font0 = ("Helvetica", 8)
    font1 = ("Helvetica", 16)
    font2 = ("Helvetica", 32)
    font3 = ("Helvetica", 64)

import tkinter as tk
# 建立視窗
win = tk.Tk()
win.title("點餐系統")
# 全螢幕
win.attributes('-fullscreen', True)     # 可以用Alt+Tab切換視窗

class unitOrderClass(object):
    def __init__(self, meals, amount, minorTotal, unitPrice, date, time):
        self.meals = meals
        self.amount = amount
        self.minorTotal = minorTotal
        self.unitPrice = unitPrice
        self.date = date
        self.time = time
    def addTreeFunction(self):
        return (self.meals, self.amount, self.minorTotal)
    def info(self):
        return (self.meals, self.amount, self.minorTotal, self.unitPrice, self.date, self.time)

class orderClass(object):
    def __init__(self, order1):
        self.order = order1

notebook = ttk.Notebook(win)     # 建立一個 Notebook , Tab 的群組
notebook.place(x=0, y=0)


# Create an instance of ttk style
style = Style()
style.theme_use('default')
style.configure('TNotebook.Tab', background="#D2B48C")
style.map("TNotebook", background=[("selected", "white")])

# create frames  第二層-經典漢堡
frame0 = ttk.Frame(notebook, width=1000, height=960)
frame0.pack(fill='both', expand=True)
# create frames  第二層-清爽摘鮮綠
frame1 = ttk.Frame(notebook, width=1000, height=960)
frame1.pack(fill='both', expand=True)
# create frames  第二層-五穀大麥堡
frame2 = ttk.Frame(notebook, width=1000, height=960)
frame2.pack(fill='both', expand=True)
# create frames  第二層-健康藜麥堡
frame3 = ttk.Frame(notebook, width=1000, height=960)
frame3.pack(fill='both', expand=True)
# create frames  第二層-精選米堡
frame4 = ttk.Frame(notebook, width=1000, height=960)
frame4.pack(fill='both', expand=True)
# create frames  第二層-經典熱狗堡
frame5 = ttk.Frame(notebook, width=1000, height=960)
frame5.pack(fill='both', expand=True)
# create frames  第二層-健康沙拉
frame6 = ttk.Frame(notebook, width=1000, height=960)
frame6.pack(fill='both', expand=True)
# create frames  第二層-精選副餐
frame7 = ttk.Frame(notebook, width=1000, height=960)
frame7.pack(fill='both', expand=True)
# create frames  第二層-嚴選茶飲
frame8 = ttk.Frame(notebook, width=1000, height=960)
frame8.pack(fill='both', expand=True)
# create frames  第二層-暖心熱茶
frame9 = ttk.Frame(notebook, width=1000, height=960)
frame9.pack(fill='both', expand=True)
# create frames  第二層-經典咖啡
frame10 = ttk.Frame(notebook, width=1000, height=960)
frame10.pack(fill='both', expand=True)
# create frames  第二層-牛奶咖啡
frame11 = ttk.Frame(notebook, width=1000, height=960)
frame11.pack(fill='both', expand=True)
# create frames  第二層-其他飲品
frame12 = ttk.Frame(notebook, width=1000, height=960)
frame12.pack(fill='both', expand=True)
# create frames  第二層-濃郁濃湯
frame13 = ttk.Frame(notebook, width=1000, height=960)
frame13.pack(fill='both', expand=True)
# create frames  第二層-幸福甜點
frame14 = ttk.Frame(notebook, width=1000, height=960)
frame14.pack(fill='both', expand=True)
# add frames to notebook
notebook.add(frame0, text='經典漢堡')
notebook.add(frame1, text='清爽摘鮮綠')
notebook.add(frame2, text='五穀大麥堡')
notebook.add(frame3, text='健康藜麥堡')
notebook.add(frame4, text='精選米堡')
notebook.add(frame5, text='經典熱狗堡')
notebook.add(frame6, text='健康沙拉')
notebook.add(frame7, text='精選副餐')
notebook.add(frame8, text='嚴選茶飲')
notebook.add(frame9, text='暖心熱茶')
notebook.add(frame10, text='經典咖啡')
notebook.add(frame11, text='牛奶咖啡')
notebook.add(frame12, text='其他飲品')
notebook.add(frame13, text='濃郁濃湯')
notebook.add(frame14, text='幸福甜點')

frameList = [frame0, frame1, frame2, frame3, frame4, frame5, frame6, frame7, frame8, frame9, frame10, frame11, frame12,
             frame13, frame14]          # 162~232


# 主餐點餐介面
menuImage = [["日式多蜜和牛堡.png", "摩力樂活蔬食堡.png", "樂活牛肉堡.png", "摩斯漢堡(原味辣味)(牛).png",
              "摩斯吉士漢堡(原味辣味)(牛).png", "蜜汁烤雞堡.png", "蜜汁烤雞起司堡.png", "摩斯鱈魚堡.png",
              "黃金炸蝦堡.png", "厚切培根牛肉堡(牛).png", '輕檸雙牛堡(牛).png'],
             ["摘鮮綠摩力樂活蔬食堡.png", "摘鮮綠樂活牛肉堡.png",  "_菜增量_摘鮮綠烤雞起司堡(使用奶油萵苣).png",
              "_菜增量_摘鮮綠炸蝦堡(使用奶油萵苣).png"], ["超級大麥元氣牛肉珍珠堡.png", "超級大麥海洋珍珠堡.png", "超級大麥薑燒珍珠堡.png",
             "超級大麥燒肉珍珠堡.png", "超級大麥杏鮑菇珍珠堡.png"],
             ["藜麥元氣牛肉珍珠堡(牛).png", "藜麥海洋珍珠堡.png", "藜麥薑燒珍珠堡.png", "藜麥燒肉珍珠堡(牛).png", "藜麥杏鮑菇珍珠堡.png"],
             ["元氣牛肉珍珠堡(牛).png", "海洋珍珠堡.png", "薑燒珍珠堡.png", "燒肉珍珠堡(牛).png", "杏鮑菇珍珠堡.png"],
             ["摩斯熱狗堡.png", "辣味吉利熱狗堡.png"], ["樂活嫩雞沙拉.png", "樂活鮮蔬沙拉.png", "夏威夷元氣鮮蔬沙拉.png",
             "雞肉地瓜總匯沙拉.png"], ["摩斯肉醬薯條(義式荷蘭醬風味).png", "黃金地瓜薯(L).png", "薯條(L).png",
             "明太子起司可樂餅.png", "北海道可樂餅.png", "方塊薯餅(3個).png", "摩斯雞塊.png", "和風炸雞(大).png",
             "和風雞腿塊.png", "雞塊薯條組.png", "法蘭克熱狗.png"], ["台東紅烏龍茶(冰).png", "荔枝柑橘蒟蒻冰茶.png", "紅心芭樂蒟蒻冰茶.png",
             "柚子冰茶.png", "摩斯玄米煎茶.png",  "冰紅茶(M).png",  "冰紅茶[無糖](M).png", "冰紅茶[無糖](L).png", "100_台灣冰綠茶[無糖](M).png",
             "100_台灣冰綠茶[無糖](L).png", "海鹽烏龍茶歐蕾(冰).png", "摩斯特調紅茶歐蕾(冰)(L).png"], ["柚子熱茶.png", "摩斯熱紅茶.png",
             "摩斯奶茶(熱)(M).png", "台東紅烏龍茶(熱).png"], ["早餐熱咖啡(M).png", "早餐熱咖啡(S).png", "澳式黑咖啡(冰)(M).png",
             "澳式黑咖啡(熱)(M).png", "摩斯咖啡(冰)(M).png", "摩斯咖啡(熱)(M).png", "摩斯咖啡(冰)(L).png", "摩斯咖啡(熱)(L).png"],
             ["拿鐵咖啡(冰)(M).png", "拿鐵咖啡(熱)(M).png", "拿鐵咖啡(冰)(L).png", "拿鐵咖啡(熱)(L).png", "卡布奇諾(熱)(M).png",
              "卡布奇諾(熱)(L).png", "摩斯馥樂白(熱)(M).png", "澳式冰歐蕾(M).png"], ["100_新鮮橘子汁.png", "摩斯柑橘綜合果汁.png",
             "摩斯蘋果汁.png", "可樂.png", "雪碧.png", "摩斯熱可可.png", "摩斯鮮乳.png", "摩斯純淨天然水.png", "摩斯純淨水.png"],
             ["玉米濃湯.png", "鮮菇濃湯.png"],
             ["蜂蜜檸檬蒟蒻.png", "葡萄蒟蒻.png", "熱帶雙果蒟蒻.png", "巴斯克乳酪蛋糕.png", "紅茶蘋果米斯球.png", "抹茶紅豆米派.png"]]
menuName = [["日式多蜜和牛堡", "摩力樂活蔬食堡", "樂活牛肉堡",  "摩斯漢堡(原味/辣味)(牛)",  "摩斯吉士漢堡(原味/辣味)(牛)",
             "蜜汁烤雞堡",  "蜜汁烤雞起司堡", "摩斯鱈魚堡",  "黃金炸蝦堡", "厚切培根牛肉堡(牛)", '輕檸雙牛堡(牛)'],
            ["摘鮮綠摩力樂活蔬食堡", "摘鮮綠樂活牛肉堡",  "'菜增量'摘鮮綠烤雞起司堡(使用奶油萵)", "'菜增量'摘鮮綠炸蝦堡(使用奶油萵苣)"],
            ["超級大麥元氣牛肉珍珠堡", "超級大麥海洋珍珠堡", "超級大麥薑燒珍珠堡", "超級大麥燒肉珍珠堡", "超級大麥杏鮑菇珍珠堡"],
            ["藜麥元氣牛肉珍珠堡(牛)", "藜麥海洋珍珠堡", "藜麥薑燒珍珠堡", "藜麥燒肉珍珠堡(牛)", "藜麥杏鮑菇珍珠堡"],
            ["元氣牛肉珍珠堡(牛)", "海洋珍珠堡", "薑燒珍珠堡", "燒肉珍珠堡(牛)", "杏鮑菇珍珠堡"], ["摩斯熱狗堡", "辣味吉利熱狗堡"],
            ["樂活嫩雞沙拉", "樂活鮮蔬沙拉", "夏威夷元氣鮮蔬沙拉", "雞肉地瓜總匯沙拉"], ["摩斯肉醬薯條(義式荷蘭醬風味)", "黃金地瓜薯(L)",
            "薯條(L)", "明太子起司可樂餅", "北海道可樂餅", "方塊薯餅(3個)", "摩斯雞塊", "和風炸雞(大)",
            "和風雞腿塊", "雞塊薯條組", "法蘭克熱狗"], ["台東紅烏龍茶(冰)", "荔枝柑橘蒟蒻冰茶", "紅心芭樂蒟蒻冰茶", "柚子冰茶",
            "摩斯玄米煎茶",  "冰紅茶(M)",  "冰紅茶[無糖](M)", "冰紅茶[無糖](L)", "100%台灣冰綠茶[無糖](M)", "100%台灣冰綠茶[無糖](L)",
            "海鹽烏龍茶歐蕾(冰)", "摩斯特調紅茶歐蕾(冰)(L)"], ["柚子熱茶", "摩斯熱紅茶", "摩斯奶茶(熱)(M)", "台東紅烏龍茶(熱)"],
            ["早餐熱咖啡(M)", "早餐熱咖啡(S)", "澳式黑咖啡(冰)(M)", "澳式黑咖啡(熱)(M)", "摩斯咖啡(冰)(M)", "摩斯咖啡(熱)(M)",
             "摩斯咖啡(冰)(L)", "摩斯咖啡(熱)(L)"], ["拿鐵咖啡(冰)(M)", "拿鐵咖啡(熱)(M)", "拿鐵咖啡(冰)(L)", "拿鐵咖啡(熱)(L)",
             "卡布奇諾(熱)(M)", "卡布奇諾(熱)(L)", "摩斯馥樂白(熱)(M)", "澳式冰歐蕾(M)"], ["100%新鮮橘子汁", "摩斯柑橘綜合果汁",
             "摩斯蘋果汁", "可樂", "雪碧", "摩斯熱可可", "摩斯鮮乳", "摩斯純淨天然水", "摩斯純淨水"], ["玉米濃湯", "鮮菇濃湯"],
            ["蜂蜜檸檬蒟蒻", "葡萄蒟蒻", "熱帶雙果蒟蒻", "巴斯克乳酪蛋糕", "紅茶蘋果米斯球", "抹茶紅豆米派"]]
menuPrice = [[99, 110, 80, 70, 75, 70, 75, 70, 75, 100, 100], [115, 85, 80, 80], [115, 85, 75, 80, 80],
             [110, 80, 70, 75, 75], [105, 75, 65, 70, 70], [55, 70], [89, 65, 50, 65], [75, 55, 45, 60, 35, 35, 65, 85,
             65, 60, 35], [60, 80, 75, 75, 45, 35, 35, 40, 35, 40, 85, 75], [50, 40, 60, 60], [40, 35, 65, 65, 50, 50,
             60, 60], [70, 70, 80, 80, 70, 80, 85, 85], [80, 35, 35, 35, 35, 45, 35, 20, 16], [40, 40], [40, 40, 45,
             85, 50, 45]]
menus = ["hamburger", "vegetable", "riceBurger_Barley", "riceBurger_Quinoa", "riceBurger_Rice", "hotDogHan", "salad",
         "dessert", "iceTea", "hotTea", "coffee", "milkCoffee", "otherDrink", "soup", "sweet"]
menuLabel = []
buttons = []
img = []

imgB = []
labelImgB = []
checkBotton = []
delBotton = []
labelTotalNum = []
tree = []
# 利用迴圈將背景圖放上去每個frame
for a in range(len(frameList)):
    # 放置背景圖
    imgB.append("imgB"+str(a))
    labelImgB.append("labelImgB"+str(a))
    imgB[a] = ImageTk.PhotoImage(Image.open("white.png"))
    labelImgB[a] = tk.Label(frameList[a], image=imgB[a])
    labelImgB[a].place(x=0, y=0)

def back():
    global newWin
    newWin.destroy()


counters = 0
def checkoutFunc():
    global newWin
    total = 0
    try:
        df = pd.read_excel('/Users/zhangyouxuan/Documents/面試作品/POS/POS.xlsx')
        num = df.sort_values(by=['訂單編號'], ascending=False)[:1]['訂單編號']
        ID = int(num) + 1
    except:
        ID = int(dt.datetime.today().strftime("%Y%m%d")) * 1000 + 1
    for i in orderTree.get_children():
        item = orderTree.item(i)
        record = item['values']
        name = record[0]
        quantity = record[1]
        unitPrice = record[3]
        subtotal = record[2]
        date = record[4]
        time1 = record[5]
        total = total + subtotal
        writeOrderToExcel(ID, name, quantity, unitPrice, subtotal, date, time1)
    writeTotalToExcel(total)
    # create new windows
    newWin = Toplevel()
    newWin.title("結帳")
    zero()
    finish()


def zero():
    global total
    global labelCheck
    global labTotal
    global spin_box
    # clear all
    orderTree.delete(*orderTree.get_children())
    labelCheck['text'] = ''
    labTotal['text'] = ''
    total = 0
    spin_box.set(1)

def finish():
    global newWin
    global counters
    # count
    if counters > 99:
        counters = 1
    else:
        counters = counters + 1
    # create label and buttons
    labelOnFin = tk.Label(newWin, text="您的取餐編號為 %d\n\n請稍候等待叫號取餐\n\n感謝您的光臨" % counters,
                          bg="#ececec", fg="gray", font=font3)
    labelOnFin.place(x=430, y=200)
    buttonOnFin = tk.Button(newWin, text="返回點餐", command=back, bg="#233142", fg="gray", font=font2)
    buttonOnFin.place(x=620, y=620)

def writeTotalToExcel(total):
    global wb
    wb = load_workbook('/Users/zhangyouxuan/Documents/面試作品/POS/POS.xlsx')  # 讀取檔案
    sheet = wb.active  # 打開一個工作欄
    sheet.cell(row=sheet.max_row, column=8).value = total
    wb.save('/Users/zhangyouxuan/Documents/面試作品/POS/POS.xlsx')

def writeOrderToExcel(ID, name, quantity, unitPrice, subtotal, date, time1):
    global wb
    wb = load_workbook('/Users/zhangyouxuan/Documents/面試作品/POS/POS.xlsx')  # 讀取檔案
    sheet = wb.active  # 打開一個工作欄
    t1 = sheet.max_row + 1
    sheet.cell(row=t1, column=1).value = ID
    sheet.cell(row=t1, column=2).value = name
    sheet.cell(row=t1, column=3).value = quantity
    sheet.cell(row=t1, column=4).value = unitPrice
    sheet.cell(row=t1, column=5).value = subtotal
    sheet.cell(row=t1, column=6).value = date
    sheet.cell(row=t1, column=7).value = time1
    wb.save('/Users/zhangyouxuan/Documents/面試作品/POS/POS.xlsx')


if not os.path.exists('/Users/zhangyouxuan/Documents/面試作品/POS/POS.xlsx'):
    data = pd.ExcelWriter("/Users/zhangyouxuan/Documents/面試作品/POS/POS.xlsx", engine='xlsxwriter')
    df = pd.DataFrame({'訂單編號': [], '品名': [], '數量': [], '單價': [],
                       '小計': [], '銷貨日期': [], '銷貨時間': [], '總計': []})
    df.to_excel(data, sheet_name='Sheet1', index=False)
    data.save()
else:
    print("")


# add tree

def addTree():
    global price
    global order
    global spin_box
    global total
    global labTotal
    global labelCheck
    # update total
    total = 0
    subtotal = price * int(spin_box.get())
    # add to tree
    data = unitOrderClass(order, spin_box.get(), subtotal, price, dt.date.today(), dt.datetime.today().strftime("%H:%M:%S"))
    orderTree.insert('', tk.END, values=data.info())
    # change total on win
    for i in orderTree.get_children():
        item = orderTree.item(i)
        record = item['values']
        total = total + record[2]
    labTotal['text'] = str(total)
    # clear windows
    labelCheck['text'] = ''
    spin_box.set(1)

# delete tree
def delTree():
    total = 0
    for selected_item in orderTree.selection():
        orderTree.delete(selected_item)
    for i in orderTree.get_children():
        item = orderTree.item(i)
        record = item['values']
        total = total + record[2]
    labTotal['text'] = str(total)




# 點餐按鈕連動function
price = 0
order = ""
# 按下餐點圖片按鈕後連結function
def orderFunc(j, i):
    global order
    global price
    global spin_box
    global labTotal
    order = menuName[j][i]       # 421
    price = menuPrice[j][i]
    for b in range(len(frameList)):
        labelCheck['text'] = order

"""
def loging():
    global entryAccountString
    global entryPasswordString
    global adminWin
    if entryAccountString.get() == "admin" and (entryPasswordString.get() == "admin"):
        adminWin.destroy()
        settingWin = Toplevel()
        setting.user(settingWin)


def adminFunc():
    global entryAccountString
    global entryPasswordString
    global adminWin
    adminWin = Toplevel()
    # account
    labelAccountTitle = tk.Label(adminWin, text="Please enter your account:")
    labelAccountTitle.place(x=570, y=300)
    entryAccountString = tk.StringVar()
    entryAccount = tk.Entry(adminWin, textvariable=entryAccountString, width=12, justify=RIGHT)
    entryAccount.place(x=770, y=300)
    # password
    labelPasswordTitle = tk.Label(adminWin, text="Please enter your password:")
    labelPasswordTitle.place(x=570, y=350)
    entryPasswordString = tk.StringVar()
    entryPassword = tk.Entry(adminWin, textvariable=entryPasswordString, width=12, justify=RIGHT)
    entryPassword.place(x=770, y=350)
    # check
    logingButton = tk.Button(adminWin, text="Loging", bg="#ececec", font=font1, command=loging)
    logingButton.place(x=800, y=400)
"""
# create item on the window（按下餐點按鈕後顯示品名）
labelCheck = tk.Label(win, text='', bg="#ececec", font=font1)
labelCheck.place(x=1050, y=20)
# create spin_box on the window
current_value = tk.StringVar(value=1)
spin_box = ttk.Spinbox(win, from_=1, to=30, textvariable=current_value)
spin_box.place(x=1050, y=50)
# create add  button on the window（
addButton = tk.Button(win, text="新增", bg="#ececec", font=font1, command=addTree)
addButton.place(x=1300, y=100)
# create delete button on the window
deleteButton = tk.Button(win, text="刪除", bg="#ececec", font=font1, command=delTree)
deleteButton.place(x=1200, y=100)
# create checkout button on the window
checkoutButton = tk.Button(win, text="結帳", bg="#ececec", font=font2, command=checkoutFunc)
checkoutButton.place(x=1300, y=800)
# label total title
labTotalTitle = tk.Label(win, text="Total", font=font2).place(x=1050, y=750)
# label total
labTotal = tk.Label(win, text='', font=font2)
labTotal.place(x=1250, y=750)
"""
# create button on the windows
adminButton = tk.Button(win, text="Administrator", bg="#ececec", font=font0, command=adminFunc)
adminButton.place(x=1360, y=875)
"""



# choose tree function
def Goodselected(event):
    for sale_item in orderTree.selection():
        item = orderTree.item(sale_item)  # 被選的項目資料 Dict
        record = item['values']
        labelCheck['text'] = record[0]
        spin_box.set(record[1])


# create tree
columns = ('餐點品項', '數量', '價格', '單價', '日期', '時間')
orderTree = ttk.Treeview(win, columns=columns, show='headings', height=20)          # 設定Treeview欄位名稱
style = ttk.Style()
style.configure("Treeview.Heading", font=(None, 20))
orderTree.column('餐點品項', width=200, anchor=tk.E)
orderTree.column('數量', width=80, anchor=tk.E)
orderTree.column('價格', width=115, anchor=tk.E)
orderTree.heading('餐點品項', text='餐點品項')  # 欄位文字設定
orderTree.heading('數量', text='數量')  # 欄位文字設定
orderTree.heading('價格', text='價格')  # 欄位文字設定
orderTree.bind('<<TreeviewSelect>>', Goodselected)  # 綁定事件 選取時
orderTree.place(x=1043, y=300)










# 設置迴圈 第一層為餐單種類 第二層為各種類菜單品項
for j in range(len(menuImage)):
    img.append([])
    buttons.append([])
    menuLabel.append([])
    for i in range(len(menuImage[j])):
        digit = i // 4
        n = i % 4
        image = Image.open(menuImage[j][i])
        zoom = 0.26
        buttons[j].append(menus[j] + "Button" + str(i))
        menuLabel[j].append(menus[j] + str(i))
        img[j].append("img_" + str(j) + "_" + str(i))
        # multiple image size by zoom
        pixels_x, pixels_y = tuple([int(zoom * x) for x in image.size])
        img[j][i] = ImageTk.PhotoImage(image.resize((pixels_x, pixels_y)))
        buttons[j][i] = tk.Button(frameList[j], text=menuLabel[j][i],
                                  command=lambda idj=j, idi=i: orderFunc(idj, idi), image=img[j][i], bd=0)    # 387
        buttons[j][i].place(x=n * 265, y=digit * 290)
        menuLabel[j][i] = tk.Label(frameList[j], text="%s\n$ %d" % (menuName[j][i], menuPrice[j][i]),
                                   bg="white", fg="gray", font=font1, width=29)    # 先把文字底色顯示為黃色檢查圖片按鈕寬度字元
        menuLabel[j][i].place(x=n * 265, y=digit * 290 + 245)


win.mainloop()
